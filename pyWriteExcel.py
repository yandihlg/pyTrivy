import openpyxl

def write_excel_file(data_map, file_path):
    # Crear un nuevo archivo de Excel
    workbook = openpyxl.Workbook()

    # Seleccionar la hoja de trabajo
    worksheet = workbook.active

    # Escribir los encabezados de las columnas
    worksheet.cell(row=1, column=1, value="Imagen")
    worksheet.cell(row=2, column=1, value="Tag")
    worksheet.cell(row=3, column=1, value="Vulnerabilidades")

    # Iterar sobre el mapa y escribir los datos en el archivo de Excel
    row = 2
    for key, value in data_map.items():
        worksheet.cell(row=row, column=1, value=key)
        worksheet.cell(row=row, column=2, value=value.tag)
        worksheet.cell(row=row, column=3, value=", ".join(value.vulnerabilities))
        row += 1

    # Guardar el archivo de Excel
    workbook.save(file_path)
